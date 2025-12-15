import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors


class ReportGenerator:
    def __init__(self):
        print("📊 Report Generator Inizializzato")

    def _format_composer(self, comp_text):
        """Formatta il compositore: se vuoto/sconosciuto -> 'Non rilevato'"""
        if not comp_text: return "Non rilevato"
        t = str(comp_text).strip().lower()
        if t in ["sconosciuto", "ricerca...", "—", "-"]:
            return "Non rilevato"
        return str(comp_text).strip()

    # --- 1. EXCEL (OFFICIAL - MODIFICATO) ---
    def generate_excel(self, playlist, metadata=None):
        wb = Workbook()
        ws = wb.active
        ws.title = "Programma Musicale"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid"
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = ["N.", "TITOLO OPERA", "COMPOSITORE / AUTORE", "ARTISTA ESECUTORE"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        row_num = 2
        index_display = 1
        
        for song in playlist:
            # [MODIFICA] Salta brani cancellati o non confermati
            if song.get('is_deleted', False):
                continue
            if not song.get("confirmed", False):
                continue

            # [MODIFICA] Usa i dati MODIFICATI (title, artist, composer)
            title = song.get("title", "").strip().upper()
            artist = song.get("artist", "").strip().title()
            
            # [MODIFICA] Logica "Non rilevato" per Compositore
            raw_comp = song.get("composer", "")
            composer = self._format_composer(raw_comp).upper()

            ws.cell(row=row_num, column=1, value=index_display).alignment = center_align
            ws.cell(row=row_num, column=2, value=title).alignment = left_align
            ws.cell(row=row_num, column=3, value=composer).alignment = left_align
            ws.cell(row=row_num, column=4, value=artist).alignment = left_align

            for c in range(1, 5):
                ws.cell(row=row_num, column=c).border = thin_border

            row_num += 1
            index_display += 1

        ws.column_dimensions["A"].width = 5
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 30

        footer_row = row_num + 2
        info_text = f"Generato automaticamente il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}"
        if metadata and "artist" in metadata:
            info_text += f" - Evento: {metadata['artist']}"
        ws.cell(row=footer_row, column=2, value=info_text).font = Font(
            italic=True, size=9, color="555555"
        )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    # --- 2. PDF OFFICIAL (MODIFICATO) ---
    def generate_pdf_official(self, playlist, metadata=None):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("Programma Musicale - Borderò ufficiale", styles["Title"]))
        subtitle = "Elenco dei brani confermati."
        if metadata and "artist" in metadata:
            subtitle += f" Evento: {metadata['artist']}."
        story.append(Paragraph(subtitle, styles["Normal"]))
        story.append(Spacer(1, 12))

        header_style = ParagraphStyle(
            "HeaderStyle",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=9,
            textColor=colors.white,
            alignment=1,
        )

        data = [[
            Paragraph("N.", header_style),
            Paragraph("Titolo opera", header_style),
            Paragraph("Compositore / Autore", header_style),
            Paragraph("Artista esecutore", header_style),
        ]]

        cell_style = ParagraphStyle("CellStyle", parent=styles["Normal"], fontSize=9, leading=11)
        center_style = ParagraphStyle("CenterStyle", parent=styles["Normal"], fontSize=9, alignment=1)

        idx = 1
        for song in playlist:
            # [MODIFICA] Filtra cancellati e non confermati
            if song.get('is_deleted', False):
                continue
            if not song.get("confirmed", False):
                continue

            # [MODIFICA] Usa dati modificati
            title = (song.get("title") or "").strip()
            artist = (song.get("artist") or "").strip()
            
            raw_comp = song.get("composer")
            composer = self._format_composer(raw_comp)

            data.append([
                Paragraph(str(idx), center_style),
                Paragraph(title, cell_style),
                Paragraph(composer, cell_style),
                Paragraph(artist, cell_style),
            ])
            idx += 1

        if len(data) == 1:
            data.append(["—", "Nessun brano confermato", "", ""])

        table = Table(data, colWidths=[30, 205, 170, 120], repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))

        story.append(table)
        story.append(Spacer(1, 12))

        info_text = f"Generato automaticamente il {datetime.now().strftime('%d/%m/%Y alle %H:%M')}"
        if metadata and "artist" in metadata:
            info_text += f" - Evento: {metadata['artist']}"
        story.append(Paragraph(info_text, styles["Italic"]))

        doc.build(story)
        buffer.seek(0)
        return buffer

    # --- 3. PDF RAW (LOG TECNICO - ORIGINALE + CANCELLATI) ---
    def generate_pdf_raw(self, playlist, metadata=None):
        """
        PDF 'Log Tecnico':
        - Include TUTTI i brani (anche deleted).
        - Usa SOLO dati ORIGINALI (non modificati in Page3).
        - Deleted -> Testo Rosso.
        - Manual -> "Inserimento Manuale".
        - Sconosciuto -> "Non rilevato".
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("Log di Rilevamento Automatico", styles["Title"]))

        subtitle_text = (
            "Report tecnico di legittimità. I dati riportati corrispondono a quanto rilevato "
            "dall'algoritmo, escludendo modifiche manuali."
        )
        story.append(Paragraph(subtitle_text, styles["Normal"]))
        story.append(Spacer(1, 12))

        # Stile Header
        h_style = ParagraphStyle(
            "RawHeader",
            parent=styles["Normal"],
            fontName="Courier-Bold",
            fontSize=8,
            textColor=colors.white,
            alignment=1,  # Center
        )

        # Stile Celle Standard
        c_style = ParagraphStyle(
            "RawCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10,
            alignment=0,  # Left
        )

        # Stile ID
        c_center = ParagraphStyle(
            "RawCellCenter",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            alignment=1,
        )

        data = [[
            Paragraph("ID", h_style),
            Paragraph("Titolo (Rilevato)", h_style),
            Paragraph("Compositore (Rilevato)", h_style),
            Paragraph("Artista (Rilevato)", h_style),
        ]]

        # Lista di stili condizionali (per colorare di rosso i cancellati)
        table_styles = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2b2b2b")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            # Background alternato per leggibilità
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f4f4")]),
        ]

        if not playlist:
            data.append(["—", "Nessun dato", "", ""])
        else:
            for i, song in enumerate(playlist):
                # Calcola indice riga (header è 0, prima song è 1)
                row_idx = i + 1
                
                is_manual = song.get("manual", False)
                is_deleted = song.get("is_deleted", False)

                # --- LOGICA INSERIMENTO MANUALE ---
                if is_manual:
                    song_id = "—"
                    display_title = "(Inserimento Manuale)"
                    display_comp = "(Inserimento Manuale)"
                    display_art = "—"

                # --- LOGICA RILEVAMENTO AUTOMATICO ---
                else:
                    song_id = str(song.get("id", "?"))

                    # Recupero dati ORIGINALI (non modificati)
                    orig_title = song.get("original_title")
                    orig_comp = song.get("original_composer")
                    orig_art = song.get("original_artist")

                    # Titolo
                    display_title = orig_title if orig_title else "(Dati mancanti)"

                    # Compositore: Normalizzazione "Non rilevato"
                    display_comp = self._format_composer(orig_comp)

                    # Artista
                    display_art = orig_art if orig_art else "—"

                # Creazione Paragrafi con Colore Condizionale
                # Se è cancellato, forza il colore rosso direttamente nello stile del paragrafo
                # (ReportLab Table Style TEXTCOLOR a volte fa i capricci con Paragraphs, meglio stile diretto)
                
                curr_color = colors.red if is_deleted else colors.black
                
                p_id = Paragraph(song_id, ParagraphStyle("tmp_id", parent=c_center, textColor=curr_color))
                p_tit = Paragraph(display_title, ParagraphStyle("tmp_tit", parent=c_style, textColor=curr_color))
                p_comp = Paragraph(display_comp, ParagraphStyle("tmp_comp", parent=c_style, textColor=curr_color))
                p_art = Paragraph(display_art, ParagraphStyle("tmp_art", parent=c_style, textColor=curr_color))

                data.append([p_id, p_tit, p_comp, p_art])

        col_widths = [30, 190, 190, 130]
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle(table_styles))

        story.append(table)
        story.append(Spacer(1, 12))

        info_text = f"Snapshot DB generato il {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
        if metadata and "artist" in metadata:
            info_text += f" | Session: {metadata['artist']}"
        story.append(Paragraph(info_text, styles["Italic"]))

        story.append(Spacer(1, 4))
        story.append(Paragraph(
            "Questo documento certifica l'output del sistema di riconoscimento automatico.",
            styles["Italic"],
        ))

        doc.build(story)
        buffer.seek(0)
        return buffer